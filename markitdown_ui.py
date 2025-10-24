import sys
from pathlib import Path
import warnings
import re
from PySide6.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout,
                              QHBoxLayout, QLabel, QLineEdit, QPushButton,
                              QTextEdit, QFileDialog, QMessageBox, QProgressBar,
                              QListWidget, QListWidgetItem, QFrame,
                              QAbstractItemView)
from PySide6.QtCore import QThread, Signal, Qt
from PySide6.QtGui import QFont, QDragEnterEvent, QDropEvent

try:
    import openpyxl
    EXCEL_SUPPORT = True
except ImportError:
    EXCEL_SUPPORT = False

# 忽略各种警告
warnings.filterwarnings("ignore", message="Couldn't find ffmpeg or avconv")
warnings.filterwarnings("ignore", message="Unsupported Windows version")
warnings.filterwarnings("ignore", category=UserWarning, module="onnxruntime")

try:
    from markitdown import MarkItDown, UnsupportedFormatException, MissingDependencyException
except ImportError as e:
    print("Error: Cannot import markitdown library")
    print("Please run: pip install markitdown[all]")
    print(f"Details: {e}")
    sys.exit(1)

# 转换工作线程
class ConversionWorker(QThread):
    finished = Signal(str, str)  # markdown_content, source
    error = Signal(str)
    
    def __init__(self, md_instance, source, excel_file=None, selected_sheets=None):
        super().__init__()
        self.md = md_instance
        self.source = source
        self.excel_file = excel_file
        self.selected_sheets = selected_sheets
    
    def run(self):
        try:
            # 检查是否为 Excel 文件且需要特殊处理
            if (self.excel_file and 
                self.excel_file == self.source and 
                EXCEL_SUPPORT and 
                self.selected_sheets):
                
                # 使用自定义的 Excel 转换
                markdown_content = self._convert_excel_sheets(self.source, self.selected_sheets)
                self.finished.emit(markdown_content, self.source)
            else:
                # 使用 MarkItDown 的默认转换
                result = self.md.convert(self.source)
                self.finished.emit(result.markdown, self.source)
                
        except UnsupportedFormatException:
            self.error.emit("不支持的文件格式")
        except MissingDependencyException as e:
            self.error.emit(f"缺少依赖: {e}")
        except Exception as e:
            self.error.emit(f"转换失败: {str(e)}")
    
    def _convert_excel_sheets(self, filename, selected_sheets):
        """转换选中的 Excel sheets"""
        if not selected_sheets:
            raise Exception("请至少选择一个 Sheet")
        
        results = []
        for sheet_name in selected_sheets:
            try:
                workbook = openpyxl.load_workbook(filename, read_only=True)
                worksheet = workbook[sheet_name]
                
                # 将 sheet 数据转换为 markdown 表格
                markdown_content = self._worksheet_to_markdown(worksheet, sheet_name)
                results.append(markdown_content)
                
                workbook.close()
                
            except Exception as e:
                results.append(f"# {sheet_name}\n\n**错误**: 无法转换此 Sheet - {str(e)}\n\n")
        
        return "\n\n---\n\n".join(results)
    
    def _worksheet_to_markdown(self, worksheet, sheet_name):
        """将 Excel worksheet 转换为 Markdown"""
        markdown = f"# {sheet_name}\n\n"
        
        # 获取有数据的区域
        if worksheet.max_row == 1 and worksheet.max_column == 1:
            return markdown + "此 Sheet 为空\n"
        
        # 转换为表格
        rows = []
        for row in worksheet.iter_rows(values_only=True):
            # 跳过完全空的行
            if all(cell is None or str(cell).strip() == '' for cell in row):
                continue
            # 将 None 值转换为空字符串，其他值转换为字符串
            row_data = [str(cell) if cell is not None else '' for cell in row]
            rows.append(row_data)
        
        if not rows:
            return markdown + "此 Sheet 为空\n"
        
        # 确定最大列数
        max_cols = max(len(row) for row in rows) if rows else 0
        
        # 补齐所有行到相同列数
        for row in rows:
            while len(row) < max_cols:
                row.append('')
        
        # 生成 Markdown 表格
        if rows:
            # 表头
            header = "| " + " | ".join(rows[0]) + " |"
            separator = "| " + " | ".join(['---'] * len(rows[0])) + " |"
            markdown += header + "\n" + separator + "\n"
            
            # 数据行
            for row in rows[1:]:
                markdown += "| " + " | ".join(row) + " |\n"
        
        return markdown


# 支持拖拽的文本编辑器
class DragDropTextEdit(QTextEdit):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setAcceptDrops(True)
        
    def dragEnterEvent(self, event: QDragEnterEvent):
        if event.mimeData().hasUrls():
            event.acceptProposedAction()
        else:
            super().dragEnterEvent(event)
    
    def dropEvent(self, event: QDropEvent):
        if event.mimeData().hasUrls():
            urls = event.mimeData().urls()
            if urls:
                file_path = urls[0].toLocalFile()
                if file_path:
                    # 发送信号给主窗口
                    main_window = self.window()
                    if hasattr(main_window, 'handle_file_drop'):
                        main_window.handle_file_drop(file_path)
            event.acceptProposedAction()
        else:
            super().dropEvent(event)


# 支持拖拽的输入框
class DragDropLineEdit(QLineEdit):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setAcceptDrops(True)
        
    def dragEnterEvent(self, event: QDragEnterEvent):
        if event.mimeData().hasUrls():
            event.acceptProposedAction()
        else:
            super().dragEnterEvent(event)
    
    def dropEvent(self, event: QDropEvent):
        if event.mimeData().hasUrls():
            urls = event.mimeData().urls()
            if urls:
                file_path = urls[0].toLocalFile()
                if file_path:
                    self.setText(file_path)
                    # 通知主窗口文件已更改
                    main_window = self.window()
                    if hasattr(main_window, 'handle_file_drop'):
                        main_window.handle_file_drop(file_path)
            event.acceptProposedAction()
        else:
            super().dropEvent(event)


class MarkItDownUI(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("MarkItDown 文件转换器")
        self.setGeometry(100, 100, 1000, 850)
        self.setMinimumSize(850, 700)

        # 初始化变量
        self.excel_sheets = []
        self.selected_sheets = []
        self.current_excel_file = None
        self.current_result = ""
        self.current_title = ""

        # 设置现代化样式
        self.setup_style()

        # 初始化MarkItDown
        try:
            self.md = MarkItDown()
            self.setup_ui()
        except Exception as e:
            QMessageBox.critical(self, "初始化错误", f"无法初始化MarkItDown: {e}")
            sys.exit(1)

    def setup_style(self):
        """设置现代化的应用样式 - 基于Material Design原则"""
        self.setStyleSheet("""
            /* ===== 主窗口 ===== */
            QMainWindow {
                background-color: #f8f9fa;
            }

            /* ===== 卡片容器 ===== */
            QWidget#cardContainer {
                background-color: white;
                border-radius: 12px;
                border: 1px solid #e9ecef;
            }

            /* ===== 分组框 ===== */
            QGroupBox {
                background-color: white;
                border: none;
                border-radius: 12px;
                margin-top: 8px;
                padding: 20px 16px 16px 16px;
                font-weight: 600;
                font-size: 14px;
                color: #212529;
            }

            QGroupBox::title {
                subcontrol-origin: margin;
                subcontrol-position: top left;
                left: 16px;
                top: 8px;
                padding: 0 8px;
                background-color: white;
                color: #495057;
            }

            /* ===== 输入框 ===== */
            QLineEdit {
                padding: 0px 12px;
                border: 2px solid #dee2e6;
                border-radius: 6px;
                background-color: #ffffff;
                font-size: 12px;
                color: #212529;
                selection-background-color: #0d6efd;
                selection-color: white;
                min-height: 20px;
            }

            QLineEdit:hover {
                border: 2px solid #adb5bd;
                background-color: #f8f9fa;
            }

            QLineEdit:focus {
                border: 2px solid #0d6efd;
                background-color: white;
                outline: none;
            }

            QLineEdit::placeholder {
                color: #adb5bd;
            }

            /* ===== 按钮 ===== */
            QPushButton {
                background-color: #0d6efd;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 8px 16px;
                font-size: 13px;
                font-weight: 600;
                min-width: 80px;
                min-height: 36px;
            }

            QPushButton:hover {
                background-color: #0b5ed7;
            }

            QPushButton:pressed {
                background-color: #0a58ca;
                padding: 9px 15px 7px 17px;
            }

            QPushButton:disabled {
                background-color: #e9ecef;
                color: #adb5bd;
            }

            /* 次要按钮 */
            QPushButton#secondaryButton {
                background-color: #6c757d;
                color: white;
            }

            QPushButton#secondaryButton:hover {
                background-color: #5c636a;
            }

            QPushButton#secondaryButton:pressed {
                background-color: #565e64;
            }

            /* 成功按钮 */
            QPushButton#successButton {
                background-color: #198754;
                color: white;
            }

            QPushButton#successButton:hover {
                background-color: #157347;
            }

            QPushButton#successButton:pressed {
                background-color: #146c43;
            }

            /* 危险按钮 */
            QPushButton#dangerButton {
                background-color: #dc3545;
                color: white;
            }

            QPushButton#dangerButton:hover {
                background-color: #bb2d3b;
            }

            QPushButton#dangerButton:pressed {
                background-color: #b02a37;
            }

            /* 浏览按钮 - 轮廓样式 */
            QPushButton#browseButton {
                background-color: transparent;
                color: #0d6efd;
                border: 2px solid #0d6efd;
                border-radius: 6px;
                padding: 0px 16px;
                font-size: 12px;
                font-weight: 600;
                min-width: 70px;
                min-height: 20px;
            }

            QPushButton#browseButton:hover {
                background-color: #0d6efd;
                color: white;
            }

            QPushButton#browseButton:pressed {
                background-color: #0b5ed7;
                border-color: #0b5ed7;
            }

            /* 紧凑按钮 - 用于 Excel 操作等 */
            QPushButton#compactButton {
                background-color: #e9ecef;
                color: #495057;
                border: none;
                border-radius: 4px;
                padding: 4px 12px;
                font-size: 12px;
                font-weight: 500;
                min-width: 50px;
                min-height: 28px;
            }

            QPushButton#compactButton:hover {
                background-color: #dee2e6;
                color: #212529;
            }

            QPushButton#compactButton:pressed {
                background-color: #ced4da;
                padding: 5px 11px 3px 13px;
            }

            /* ===== 文本编辑器 ===== */
            QTextEdit {
                border: 2px solid #dee2e6;
                border-radius: 8px;
                background-color: #ffffff;
                padding: 10px;
                font-family: 'Consolas', 'Monaco', 'Courier New', monospace;
                font-size: 11px;
                color: #212529;
                selection-background-color: #0d6efd;
                selection-color: white;
            }

            QTextEdit:focus {
                border: 2px solid #0d6efd;
            }

            /* 确保 placeholder 文本完整显示 */
            QTextEdit QAbstractScrollArea {
                padding: 0px;
            }

            /* ===== 列表控件 ===== */
            QListWidget {
                border: 2px solid #dee2e6;
                border-radius: 6px;
                background-color: white;
                padding: 6px;
                font-size: 12px;
                color: #212529;
                outline: none;
            }

            QListWidget:focus {
                border: 2px solid #0d6efd;
            }

            QListWidget::item {
                padding: 6px 10px;
                border-radius: 4px;
                margin: 1px 0;
                border: none;
            }

            QListWidget::item:hover {
                background-color: #f8f9fa;
            }

            QListWidget::item:selected {
                background-color: #0d6efd;
                color: white;
            }

            QListWidget::item:selected:hover {
                background-color: #0b5ed7;
            }

            /* ===== 进度条 ===== */
            QProgressBar {
                border: none;
                border-radius: 8px;
                text-align: center;
                background-color: #e9ecef;
                height: 8px;
                font-size: 11px;
                color: #495057;
            }

            QProgressBar::chunk {
                background-color: #0d6efd;
                border-radius: 8px;
            }

            /* ===== 状态标签 ===== */
            QLabel#statusLabel {
                background-color: #e7f1ff;
                border: none;
                border-radius: 8px;
                padding: 12px 16px;
                color: #084298;
                font-size: 13px;
                font-weight: 500;
            }

            QLabel#sectionTitle {
                font-size: 13px;
                font-weight: 600;
                color: #212529;
                padding: 4px 0;
            }

            /* ===== 滚动条 ===== */
            QScrollBar:vertical {
                background-color: #f8f9fa;
                width: 12px;
                border-radius: 6px;
            }

            QScrollBar::handle:vertical {
                background-color: #adb5bd;
                border-radius: 6px;
                min-height: 30px;
            }

            QScrollBar::handle:vertical:hover {
                background-color: #6c757d;
            }

            QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {
                height: 0px;
            }

            QScrollBar:horizontal {
                background-color: #f8f9fa;
                height: 12px;
                border-radius: 6px;
            }

            QScrollBar::handle:horizontal {
                background-color: #adb5bd;
                border-radius: 6px;
                min-width: 30px;
            }

            QScrollBar::handle:horizontal:hover {
                background-color: #6c757d;
            }

            QScrollBar::add-line:horizontal, QScrollBar::sub-line:horizontal {
                width: 0px;
            }
        """)

        
    def setup_ui(self):
        # 创建中央widget
        central_widget = QWidget()
        self.setCentralWidget(central_widget)

        # 主布局
        main_layout = QVBoxLayout(central_widget)
        main_layout.setSpacing(12)
        main_layout.setContentsMargins(16, 16, 16, 16)

        # ===== 顶部：输入区域 =====
        input_container = QWidget()
        input_container.setObjectName("cardContainer")
        input_layout = QVBoxLayout(input_container)
        input_layout.setSpacing(12)
        input_layout.setContentsMargins(16, 16, 16, 16)

        # 统一的输入区域（文件路径或URL）
        input_section = QWidget()
        input_section_layout = QVBoxLayout(input_section)
        input_section_layout.setSpacing(6)
        input_section_layout.setContentsMargins(0, 0, 0, 0)

        input_label = QLabel("文件或URL")
        input_label.setObjectName("sectionTitle")
        input_section_layout.addWidget(input_label)

        input_control_layout = QHBoxLayout()
        input_control_layout.setSpacing(10)

        self.file_entry = DragDropLineEdit()
        self.file_entry.setPlaceholderText("选择文件、拖拽文件到此处，或输入URL...")
        # 总高度 = min-height(20) + padding(0*2) + border(2*2) = 24px
        # 但为了垂直居中文字，使用稍大的高度
        self.file_entry.setFixedHeight(36)
        input_control_layout.addWidget(self.file_entry, stretch=1)

        browse_btn = QPushButton("浏览")
        browse_btn.setObjectName("browseButton")
        browse_btn.setMinimumWidth(80)
        # 设置和输入框完全相同的高度
        browse_btn.setFixedHeight(36)
        browse_btn.clicked.connect(self.browse_file)
        input_control_layout.addWidget(browse_btn)

        input_section_layout.addLayout(input_control_layout)
        input_layout.addWidget(input_section)

        main_layout.addWidget(input_container)

        # ===== Excel Sheet 选择区域（初始隐藏）=====
        self.excel_container = QWidget()
        self.excel_container.setObjectName("cardContainer")
        excel_main_layout = QVBoxLayout(self.excel_container)
        excel_main_layout.setSpacing(8)
        excel_main_layout.setContentsMargins(16, 12, 16, 12)

        # 标题和按钮在同一行
        excel_header_layout = QHBoxLayout()
        excel_header_layout.setSpacing(10)

        excel_title = QLabel("Excel Sheet 选择")
        excel_title.setObjectName("sectionTitle")
        excel_header_layout.addWidget(excel_title)

        excel_header_layout.addStretch()

        # 使用更小的按钮
        select_all_btn = QPushButton("全选")
        select_all_btn.setObjectName("compactButton")
        select_all_btn.clicked.connect(self.select_all_sheets)
        excel_header_layout.addWidget(select_all_btn)

        deselect_all_btn = QPushButton("全不选")
        deselect_all_btn.setObjectName("compactButton")
        deselect_all_btn.clicked.connect(self.deselect_all_sheets)
        excel_header_layout.addWidget(deselect_all_btn)

        invert_btn = QPushButton("反选")
        invert_btn.setObjectName("compactButton")
        invert_btn.clicked.connect(self.invert_sheet_selection)
        excel_header_layout.addWidget(invert_btn)

        excel_main_layout.addLayout(excel_header_layout)

        # Sheet 列表
        self.sheet_listbox = QListWidget()
        self.sheet_listbox.setSelectionMode(QAbstractItemView.MultiSelection)
        self.sheet_listbox.setMinimumHeight(100)
        self.sheet_listbox.setMaximumHeight(150)
        excel_main_layout.addWidget(self.sheet_listbox)

        main_layout.addWidget(self.excel_container)
        self.excel_container.hide()  # 初始隐藏

        # ===== 操作按钮区域 =====
        button_container = QWidget()
        button_main_layout = QVBoxLayout(button_container)
        button_main_layout.setSpacing(8)
        button_main_layout.setContentsMargins(0, 0, 0, 0)

        button_layout = QHBoxLayout()
        button_layout.setSpacing(8)

        # 主转换按钮 - 更突出
        convert_btn = QPushButton("转换为Markdown")
        convert_btn.setMinimumHeight(42)
        convert_btn.setMinimumWidth(140)
        convert_btn.clicked.connect(self.convert_file)
        button_layout.addWidget(convert_btn)

        # 添加弹性空间
        button_layout.addStretch()

        # 次要按钮 - 更小更紧凑
        save_btn = QPushButton("保存结果")
        save_btn.setObjectName("successButton")
        save_btn.setMinimumHeight(38)
        save_btn.setMinimumWidth(90)
        save_btn.clicked.connect(self.save_result)
        button_layout.addWidget(save_btn)

        clear_btn = QPushButton("清空")
        clear_btn.setObjectName("secondaryButton")
        clear_btn.setMinimumHeight(38)
        clear_btn.setMinimumWidth(70)
        clear_btn.clicked.connect(self.clear_result)
        button_layout.addWidget(clear_btn)

        button_main_layout.addLayout(button_layout)

        # 进度条（初始状态隐藏）
        self.progress = QProgressBar()
        self.progress.setRange(0, 0)  # 无限进度条
        self.progress.setMinimumHeight(6)
        button_main_layout.addWidget(self.progress)
        self.progress.hide()  # 初始隐藏

        main_layout.addWidget(button_container)

        # ===== 结果显示区域 =====
        result_container = QWidget()
        result_container.setObjectName("cardContainer")
        result_main_layout = QVBoxLayout(result_container)
        result_main_layout.setSpacing(10)
        result_main_layout.setContentsMargins(16, 16, 16, 16)

        result_title = QLabel("转换结果")
        result_title.setObjectName("sectionTitle")
        result_main_layout.addWidget(result_title)

        self.result_text = DragDropTextEdit()
        self.result_text.setPlaceholderText("转换结果将显示在这里...\n\n您也可以直接拖拽文件到此处进行转换。")
        self.result_text.setFont(QFont("Consolas", 10))
        self.result_text.setMinimumHeight(180)
        # 设置文档边距，避免文字被裁剪
        self.result_text.document().setDocumentMargin(5)
        result_main_layout.addWidget(self.result_text)

        main_layout.addWidget(result_container, stretch=1)

        # ===== 状态栏 =====
        self.status_label = QLabel("就绪 - 请选择文件或输入URL")
        self.status_label.setObjectName("statusLabel")
        main_layout.addWidget(self.status_label)
        
    def browse_file(self):
        filename, _ = QFileDialog.getOpenFileName(
            self,
            "选择要转换的文件",
            "",
            "所有支持的文件 (*.pdf *.docx *.pptx *.xlsx *.csv *.html *.epub *.jpg *.png);;PDF文件 (*.pdf);;Word文档 (*.docx);;PowerPoint (*.pptx);;Excel文件 (*.xlsx *.xls);;图像文件 (*.jpg *.jpeg *.png *.gif *.bmp);;所有文件 (*.*)"
        )
        if filename:
            self.file_entry.setText(filename)
            self._check_excel_file(filename)
    
    def handle_file_drop(self, file_path):
        """处理文件拖拽"""
        self.file_entry.setText(file_path)
        self._check_excel_file(file_path)
            
    def convert_file(self):
        source = self.file_entry.text().strip()

        if not source:
            QMessageBox.warning(self, "错误", "请选择文件或输入URL")
            return

        # 在后台线程中执行转换
        selected_sheets = self._get_selected_sheets() if self.current_excel_file else None
        
        self.worker = ConversionWorker(self.md, source, self.current_excel_file, selected_sheets)
        self.worker.finished.connect(self._conversion_complete)
        self.worker.error.connect(self._conversion_error)
        
        self._start_conversion()
        self.worker.start()
        
    def _start_conversion(self):
        self.progress.show()  # 显示进度条
        self.status_label.setText("正在转换...")
        self.result_text.clear()

    def _conversion_complete(self, markdown_content, source):
        self.progress.hide()  # 隐藏进度条
        self.status_label.setText(f"转换完成: {Path(source).name if not source.startswith('http') else source}")
        
        # 显示结果
        self.result_text.setPlainText(markdown_content)
        
        # 存储结果用于保存
        self.current_result = markdown_content
        
        # 根据源文件生成标题
        if source.startswith('http'):
            self.current_title = "web_content"
        else:
            # 使用原文件名（不含扩展名）作为标题
            source_path = Path(source)
            self.current_title = source_path.stem  # 文件名不含扩展名
    
    def _conversion_error(self, error_message):
        self.progress.hide()
        self.status_label.setText(f"转换失败: {error_message}")
        QMessageBox.critical(self, "转换错误", error_message)
        
    def _sanitize_filename(self, filename):
        """清理文件名中的非法字符"""
        # Windows文件名非法字符
        illegal_chars = r'[<>:"/\\|?*]'
        # 替换非法字符为下划线
        sanitized = re.sub(illegal_chars, '_', filename)
        # 移除多余的空格和点
        sanitized = sanitized.strip('. ')
        # 如果文件名为空，使用默认名称
        if not sanitized:
            sanitized = "converted_document"
        return sanitized

    def save_result(self):
        if not self.current_result:
            QMessageBox.warning(self, "警告", "没有可保存的转换结果")
            return
        
        # 清理文件名
        clean_title = self._sanitize_filename(self.current_title)
        
        filename, _ = QFileDialog.getSaveFileName(
            self,
            "保存Markdown文件",
            f"{clean_title}.md",
            "Markdown文件 (*.md);;文本文件 (*.txt);;所有文件 (*.*)"
        )
        
        if filename:
            try:
                with open(filename, 'w', encoding='utf-8') as f:
                    f.write(self.current_result)
                self.status_label.setText(f"已保存: {Path(filename).name}")
                QMessageBox.information(self, "成功", f"文件已保存到: {filename}")
            except Exception as e:
                QMessageBox.critical(self, "保存错误", f"保存文件失败: {str(e)}")
                
    def clear_result(self):
        self.result_text.clear()
        self.file_entry.clear()
        self.status_label.setText("就绪 - 请选择文件或输入URL")
        self.current_result = ""

        # 隐藏 Excel 选择区域
        self.excel_container.hide()
        self.current_excel_file = None
        self.excel_sheets = []
        self.selected_sheets = []

    def _check_excel_file(self, filename):
        """检查是否为 Excel 文件，如果是则显示 sheet 选择"""
        if not EXCEL_SUPPORT:
            return

        file_ext = Path(filename).suffix.lower()
        if file_ext in ['.xlsx', '.xls']:
            try:
                self.current_excel_file = filename
                self._load_excel_sheets(filename)
                self.excel_container.show()  # 显示 Excel 选择区域
            except Exception as e:
                QMessageBox.critical(self, "Excel 文件错误", f"无法读取 Excel 文件: {str(e)}")
        else:
            self.excel_container.hide()  # 隐藏 Excel 选择区域
            self.current_excel_file = None

    def _load_excel_sheets(self, filename):
        """加载 Excel 文件的所有 sheet"""
        try:
            workbook = openpyxl.load_workbook(filename, read_only=True)
            self.excel_sheets = workbook.sheetnames
            workbook.close()
            
            # 更新 listbox
            self.sheet_listbox.clear()
            for sheet in self.excel_sheets:
                item = QListWidgetItem(sheet)
                self.sheet_listbox.addItem(item)
            
            # 默认选择所有 sheet
            self.select_all_sheets()
            
        except Exception as e:
            raise Exception(f"读取 Excel 文件失败: {str(e)}")

    def select_all_sheets(self):
        """选择所有 sheet"""
        for i in range(self.sheet_listbox.count()):
            self.sheet_listbox.item(i).setSelected(True)

    def deselect_all_sheets(self):
        """取消选择所有 sheet"""
        for i in range(self.sheet_listbox.count()):
            self.sheet_listbox.item(i).setSelected(False)

    def invert_sheet_selection(self):
        """反选 sheet"""
        for i in range(self.sheet_listbox.count()):
            item = self.sheet_listbox.item(i)
            item.setSelected(not item.isSelected())

    def _get_selected_sheets(self):
        """获取选中的 sheet 名称列表"""
        selected_sheets = []
        for i in range(self.sheet_listbox.count()):
            item = self.sheet_listbox.item(i)
            if item.isSelected():
                selected_sheets.append(item.text())
        return selected_sheets


def main():
    app = QApplication(sys.argv)
    window = MarkItDownUI()
    window.show()
    sys.exit(app.exec())

if __name__ == "__main__":
    main()



